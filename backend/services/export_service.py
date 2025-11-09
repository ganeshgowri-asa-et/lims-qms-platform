"""
Form Export Service
Export form records to PDF and Excel formats
"""
from sqlalchemy.orm import Session
from backend.models.form import FormRecord
from backend.services.form_data_service import FormDataService
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import io


class ExportService:
    """Export form records to various formats"""

    def __init__(self, db: Session):
        self.db = db
        self.form_service = FormDataService(db)

    def export_to_pdf(
        self,
        record_id: int,
        output_path: str = None,
        include_signatures: bool = True,
        include_workflow: bool = True
    ) -> str:
        """
        Export form record to PDF

        Args:
            record_id: ID of the record to export
            output_path: Output file path (optional)
            include_signatures: Include signature images
            include_workflow: Include workflow history

        Returns:
            Path to generated PDF file
        """
        # Get record data
        record_data = self.form_service.get_record_with_template(record_id)

        # Generate output path if not provided
        if not output_path:
            output_dir = Path("/home/user/lims-qms-platform/exports/pdf")
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{record_data['record']['record_number']}.pdf"
            output_path = str(output_dir / filename)

        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        story.append(Paragraph(record_data['template']['name'], title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Record Information
        info_data = [
            ['Record Number:', record_data['record']['record_number']],
            ['Status:', record_data['record']['status'].upper()],
            ['Created:', record_data['record']['created_at']],
            ['Template Version:', record_data['template']['version']]
        ]

        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f4788')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))

        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # Form Fields by Section
        sections = self._group_fields_by_section(record_data['fields'])

        for section_name, fields in sections.items():
            # Section Header
            if section_name:
                section_style = ParagraphStyle(
                    'SectionHeader',
                    parent=styles['Heading2'],
                    fontSize=14,
                    textColor=colors.HexColor('#2c5aa0'),
                    spaceBefore=12,
                    spaceAfter=12
                )
                story.append(Paragraph(section_name, section_style))

            # Fields Table
            field_data = []
            for field in fields:
                label = field['field_label']
                if field['is_required']:
                    label += ' *'

                value = self._format_field_value(field)

                field_data.append([label, value])

            if field_data:
                field_table = Table(field_data, colWidths=[2.5 * inch, 3.5 * inch])
                field_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))

                story.append(field_table)
                story.append(Spacer(1, 0.2 * inch))

        # Workflow Information
        if include_workflow and record_data['workflow_history']:
            story.append(PageBreak())
            story.append(Paragraph('Workflow History', styles['Heading2']))
            story.append(Spacer(1, 0.2 * inch))

            workflow_data = [['Date/Time', 'Action', 'User', 'Status', 'Comments']]

            for entry in record_data['workflow_history']:
                workflow_data.append([
                    entry['transition_time'][:19],
                    entry['action'],
                    str(entry['actor_id']),
                    entry['to_status'],
                    entry['comments'] or '-'
                ])

            workflow_table = Table(workflow_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1.2*inch, 2*inch])
            workflow_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))

            story.append(workflow_table)

        # Signatures
        if include_signatures and record_data['signatures']:
            story.append(Spacer(1, 0.3 * inch))
            story.append(Paragraph('Digital Signatures', styles['Heading2']))
            story.append(Spacer(1, 0.2 * inch))

            sig_data = [['Role', 'User', 'Signed At', 'Verified']]

            for sig in record_data['signatures']:
                sig_data.append([
                    sig['role'].upper(),
                    str(sig['user_id']),
                    sig['signed_at'][:19],
                    '✓' if sig['is_verified'] else '✗'
                ])

            sig_table = Table(sig_data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 1*inch])
            sig_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))

            story.append(sig_table)

        # Build PDF
        doc.build(story)

        return output_path

    def export_to_excel(
        self,
        record_id: int,
        output_path: str = None,
        include_workflow: bool = True
    ) -> str:
        """
        Export form record to Excel

        Args:
            record_id: ID of the record to export
            output_path: Output file path (optional)
            include_workflow: Include workflow history sheet

        Returns:
            Path to generated Excel file
        """
        # Get record data
        record_data = self.form_service.get_record_with_template(record_id)

        # Generate output path if not provided
        if not output_path:
            output_dir = Path("/home/user/lims-qms-platform/exports/excel")
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{record_data['record']['record_number']}.xlsx"
            output_path = str(output_dir / filename)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Data"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        label_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        ws['A1'] = record_data['template']['name']
        ws['A1'].font = Font(bold=True, size=16, color="1F4788")
        ws.merge_cells('A1:B1')

        # Record Information
        row = 3
        ws[f'A{row}'] = 'Record Number:'
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = record_data['record']['record_number']

        row += 1
        ws[f'A{row}'] = 'Status:'
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = record_data['record']['status'].upper()

        row += 1
        ws[f'A{row}'] = 'Created:'
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = record_data['record']['created_at']

        row += 2

        # Section: Fields Header
        ws[f'A{row}'] = 'Field'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill

        ws[f'B{row}'] = 'Value'
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill

        row += 1

        # Form Fields
        sections = self._group_fields_by_section(record_data['fields'])

        for section_name, fields in sections.items():
            if section_name:
                ws[f'A{row}'] = section_name
                ws[f'A{row}'].font = Font(bold=True, size=12, color="2C5AA0")
                ws.merge_cells(f'A{row}:B{row}')
                row += 1

            for field in fields:
                label = field['field_label']
                if field['is_required']:
                    label += ' *'

                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'A{row}'].border = border

                value = self._format_field_value(field, for_excel=True)
                ws[f'B{row}'] = value
                ws[f'B{row}'].border = border

                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        # Workflow History Sheet
        if include_workflow and record_data['workflow_history']:
            ws_workflow = wb.create_sheet("Workflow History")

            # Headers
            headers = ['Date/Time', 'Action', 'From Status', 'To Status', 'Actor ID', 'Comments']
            for col, header in enumerate(headers, start=1):
                cell = ws_workflow.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Data
            for idx, entry in enumerate(record_data['workflow_history'], start=2):
                ws_workflow.cell(row=idx, column=1).value = entry['transition_time'][:19]
                ws_workflow.cell(row=idx, column=2).value = entry['action']
                ws_workflow.cell(row=idx, column=3).value = entry['from_status'] or '-'
                ws_workflow.cell(row=idx, column=4).value = entry['to_status']
                ws_workflow.cell(row=idx, column=5).value = entry['actor_id']
                ws_workflow.cell(row=idx, column=6).value = entry['comments'] or '-'

            # Adjust column widths
            ws_workflow.column_dimensions['A'].width = 20
            ws_workflow.column_dimensions['B'].width = 20
            ws_workflow.column_dimensions['C'].width = 20
            ws_workflow.column_dimensions['D'].width = 20
            ws_workflow.column_dimensions['E'].width = 12
            ws_workflow.column_dimensions['F'].width = 40

        # Save workbook
        wb.save(output_path)

        return output_path

    def export_template_to_excel(
        self,
        template_id: int,
        output_path: str = None
    ) -> str:
        """
        Export blank template to Excel for bulk data entry

        Args:
            template_id: ID of the template
            output_path: Output file path (optional)

        Returns:
            Path to generated Excel file
        """
        from backend.models.form import FormTemplate

        template = self.db.query(FormTemplate).filter(
            FormTemplate.id == template_id
        ).first()

        if not template:
            raise ValueError(f"Template {template_id} not found")

        # Generate output path if not provided
        if not output_path:
            output_dir = Path("/home/user/lims-qms-platform/exports/templates")
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{template.code}_template.xlsx"
            output_path = str(output_dir / filename)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Template"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")

        # Header
        ws['A1'] = template.name
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        row = 3

        # Column Headers
        headers = ['Field Name', 'Field Label', 'Type', 'Required', 'Help Text']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        row += 1

        # Fields
        for field in sorted(template.fields, key=lambda x: x.order):
            ws.cell(row=row, column=1).value = field.field_name
            ws.cell(row=row, column=2).value = field.field_label
            ws.cell(row=row, column=3).value = field.field_type.value
            ws.cell(row=row, column=4).value = 'Yes' if field.is_required else 'No'
            ws.cell(row=row, column=5).value = field.help_text or ''
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40

        wb.save(output_path)

        return output_path

    def _group_fields_by_section(self, fields: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """Group fields by section"""
        sections = {}

        for field in fields:
            section = field.get('section') or 'General'

            if section not in sections:
                sections[section] = []

            sections[section].append(field)

        return sections

    def _format_field_value(self, field: Dict[str, Any], for_excel: bool = False) -> str:
        """Format field value for display"""
        value = field.get('value')

        if value is None or value == '':
            return '-'

        # Handle different field types
        if field['field_type'] == 'CHECKBOX':
            return 'Yes' if value in [True, 'true', '1', 1] else 'No'

        elif field['field_type'] in ['MULTISELECT']:
            if isinstance(value, list):
                return ', '.join(str(v) for v in value)

        elif field['field_type'] == 'FILE':
            if isinstance(value, list):
                return f"{len(value)} file(s)" if not for_excel else '\n'.join(value)
            return str(value)

        return str(value)
